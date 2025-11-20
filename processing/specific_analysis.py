# -*- coding: utf-8 -*-
"""
Comparación morfométrica (volúmenes y asimetrías) contra base de referencia.

Entradas:
- stats_dir: carpeta con los .txt/.stats generados por aparcstats2table/asegstats2table y *.stats.
  Requeridos (siempre en modo --transpose):
    - lh_aparc.DKTatlas.mapped_volume_stats.txt
    - rh_aparc.DKTatlas.mapped_volume_stats.txt
    - aseg_stats_cm3.txt               ( --meas volume --scale=0.001 --transpose )
    - aseg_stats_etiv.txt              ( --meas volume --etiv   --transpose )
  Recomendados (para espesores y métricas globales):
    - lh.aparc.DKTatlas.mapped.stats
    - rh.aparc.DKTatlas.mapped.stats
    - aseg.stats

- base_control_path: Excel con hoja 'metricas' (volúmenes/espesores) y hoja 'Asimetrias' (rangos LI).

Salidas:
- df_comparacion: DataFrame con columnas:
  ['Measure:GrayVol','Volumen mm3 (sujeto)','Volrel% (sujeto)','Mediana','IC_95%_Bajo','IC_95%_Alto',
   'IC_99%_Bajo','IC_99%_Alto','Percentil (sujeto)','Flag','Z_sujeto','Dentro_de_Umbral_±3.5']

- df_asimetrias: DataFrame con columnas:
  ['Measure:GrayVol','LI% (Volrel)','Mediana','IC_95%_Bajo','IC_95%_Alto','IC_99%_Bajo',
   'IC_99%_Alto','Flag','rango normal']

Además exporta siempre un Excel con dos hojas: 'Volumenes' y 'Asimetrias',
resaltando fuera de IC95 (negrita) y fuera de IC99 (naranja).
"""

import os
import re
import warnings
import unicodedata
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Optional
from openpyxl.styles import NamedStyle, Font, PatternFill

# ------------------------------- Helpers de parsing y normalización -------------------------------
def seleccionar_base_control_especificos(edad, genero):
    """
    Selecciona automáticamente la base de datos control según la edad y el género.
    """
    # Ruta absoluta al directorio que contiene las bases de datos
    base_dir = "/home/usuario/Bibliografia/pipeline_v2/recursos/morfo_cerebral/especificos/"

    if edad <= 18:
        grupo = "18_29"
    elif edad <= 29:
        grupo = "18_29"
    elif edad <= 44:
        grupo = "30_44"
    elif edad <= 60:
        grupo = "45_60"
    else:
        grupo = "45_60"      

    genero = "femenino" if genero.lower() == "f" else "masculino"
    archivo_base = f"vol_esp_{genero}_{grupo}.xlsx"

    # Combinar base_dir con el archivo
    archivo_path = os.path.join(base_dir, archivo_base)
    
    # Verificar que el archivo existe
    if not os.path.exists(archivo_path):
        raise RuntimeError(f"No se encontró el archivo de base de control en: {archivo_path}")

    return archivo_path

def _read_transposed_series(path: str) -> pd.Series:
    """
    Lee un archivo de tabla transpuesta (ROI en filas, sujeto en una columna numérica).
    Devuelve una Serie: index = ROI, values = float del sujeto.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe archivo: {path}")
    df = pd.read_csv(path, sep=r"\s+", comment="#")
    idx = df.columns[0]
    # Tomar la primera columna numérica (sujeto)
    numcols = [c for c in df.columns[1:] if pd.api.types.is_numeric_dtype(df[c])]
    if not numcols:
        # intentar convertir forzadamente
        for c in df.columns[1:]:
            try:
                df[c] = pd.to_numeric(df[c], errors="coerce")
                if df[c].notna().any():
                    numcols = [c]
                    break
            except Exception:
                pass
    if not numcols:
        raise ValueError(f"No se encontró columna numérica de sujeto en {path}")
    return df.set_index(idx)[numcols[0]].astype(float)

def _normalize_key(s: str) -> str:
    """
    Normaliza string para matching robusto: lower, sin acentos, quita signos, compacta espacios.
    """
    if s is None:
        return ""
    s = str(s)
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")  # sin acentos
    s = s.lower()
    s = re.sub(r"[^a-z0-9%/:\-\s\(\)]", " ", s)  # preservo algunos signos útiles
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _parse_valores_muestreo_to_sorted(arr_like) -> np.ndarray:
    """Convierte '0.205732,0.207218,...' a np.array ordenado de float."""
    if arr_like is None or (isinstance(arr_like, float) and np.isnan(arr_like)):
        return np.array([], dtype=float)
    s = str(arr_like).strip()
    if not s:
        return np.array([], dtype=float)
    tokens = re.split(r"[,\s;]+", s)
    vec = pd.to_numeric(pd.Series(tokens), errors="coerce").dropna().to_numpy(dtype=float)
    return np.sort(vec) if vec.size else vec

def _percentil_lineal(sorted_vals: np.ndarray, x: float) -> float:
    """Percentil 0–100 por interpolación lineal del ECDF (extrapola en extremos)."""
    if sorted_vals.size == 0 or pd.isna(x):
        return float("nan")
    pos = np.linspace(0.0, 100.0, sorted_vals.size)
    return float(np.interp(float(x), sorted_vals, pos, left=0.0, right=100.0))

def _z_robusto(x: float, mediana: float, mad: float) -> float:
    """Z robusto = (x - mediana) / (1.4826 * MAD). Devuelve NaN si MAD=0 o faltantes."""
    if pd.isna(x) or pd.isna(mediana) or pd.isna(mad) or mad == 0:
        return float("nan")
    return (x - mediana) / (1.4826 * mad)

def _warn(msg: str):
    warnings.warn(msg)
    print(f"WARNING: {msg}")

# ------------------------------- Extracción desde stats ------------------------------------------

def _get_etiv_mm3(stats_dir: str) -> float:
    # 1) Preferir aseg.stats (valor absoluto en mm^3)
    aseg_stats = os.path.join(stats_dir, "aseg.stats")
    m = _parse_header_measures(aseg_stats)
    v = m.get("eTIV", None)
    if v is not None and np.isfinite(v) and v > 1e4:  # umbral de plausibilidad
        return float(v)

    # 2) Fallback a aseg_stats_etiv.txt SOLO si parece mm^3 (no 100)
    path = os.path.join(stats_dir, "aseg_stats_etiv.txt")
    try:
        s = _read_transposed_series(path)
        for k in s.index:
            kn = _normalize_key(k)
            if kn in ("etiv","estimated total intracranial vol","estimated total intracranial volume"):
                val = float(s[k])
                if val > 1e4:        # mm^3 plausibles
                    return val
                else:
                    _warn(f"eTIV en {path} parece porcentual ({val}); ignorado.")
    except Exception as e:
        _warn(f"No pude leer {path} para eTIV: {e}")

    _warn("No se pudo extraer eTIV en mm³; devolviendo NaN")
    return float("nan")


def _parse_header_measures(stats_path: str) -> Dict[str, float]:
    """
    Parsea líneas '# Measure ...' en archivos *.stats (aseg/aparc).
    Formato típico:
      '# Measure BrainSegVol, BrainSegVol, 1139523.000, mm^3'
      '# Measure MeanThickness, MeanThickness, 2.57, mm'
      '# Measure eTIV, eTIV, 1489012.134, mm^3'
    Toma:
      - nombre = segundo token (si existe), sino el primero
      - valor  = penúltimo token convertible a float (evita el '3' de mm^3)
    Devuelve dict {nombre: valor_float}
    """
    out: Dict[str, float] = {}
    if not os.path.exists(stats_path):
        return out

    with open(stats_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if not line.startswith("# Measure"):
                continue

            # Quitar prefijo y tokenizar por coma
            rest = line.split("# Measure", 1)[-1].strip()
            tokens = [t.strip() for t in rest.split(",") if t.strip()]
            if len(tokens) < 2:
                continue

            # Nombre: preferir segundo token (suele ser la 'MeasureName')
            name_token = tokens[1] if len(tokens) >= 2 else tokens[0]

            # Valor: buscar de derecha a izquierda el primer token convertible a float,
            # evitando así capturar el '3' de 'mm^3'
            val = None
            for tok in reversed(tokens):
                try:
                    # descartar unidades típicas que no son numéricas
                    if tok.lower() in ("mm^3", "mm3", "mm", "voxels", "units"):
                        continue
                    val = float(tok.replace(" ", ""))
                    break
                except Exception:
                    continue

            if val is None:
                continue

            out[name_token] = val

    return out

def _get_thickness_mm(stats_dir: str) -> Tuple[float, float, float]:
    """
    Espesor cortical LH, RH y promedio (mm) desde lh/rh.aparc.DKTatlas.mapped.stats (# Measure MeanThickness).
    Si no existen, intenta *_thickness_stats.txt (clave *_MeanThickness_thickness).
    """
    lh_path = os.path.join(stats_dir, "lh.aparc.DKTatlas.mapped.stats")
    rh_path = os.path.join(stats_dir, "rh.aparc.DKTatlas.mapped.stats")

    lh, rh = float("nan"), float("nan")

    if os.path.exists(lh_path):
        m = _parse_header_measures(lh_path)
        # clave suele ser 'MeanThickness'
        if "MeanThickness" in m:
            lh = float(m["MeanThickness"])
    if os.path.exists(rh_path):
        m = _parse_header_measures(rh_path)
        if "MeanThickness" in m:
            rh = float(m["MeanThickness"])

    # fallback a *_thickness_stats.txt
    if (not np.isfinite(lh)) or (not np.isfinite(rh)):
        for hemi, var in (("lh", "lh"), ("rh", "rh")):
            path = os.path.join(stats_dir, f"{hemi}_aparc.DKTatlas.mapped_thickness_stats.txt")
            if not os.path.exists(path):
                continue
            s = _read_transposed_series(path)
            # buscar clave *_MeanThickness_thickness
            cand = [k for k in s.index if _normalize_key(k).endswith("meanthickness_thickness") and k.lower().startswith(hemi)]
            if cand:
                if hemi == "lh":
                    lh = float(s[cand[0]])
                else:
                    rh = float(s[cand[0]])

    prom = (lh + rh) / 2.0 if np.isfinite(lh) and np.isfinite(rh) else float("nan")
    return lh, rh, prom

def _get_global_measures_mm3(stats_dir: str) -> Dict[str, float]:
    """
    Métricas globales desde aseg.stats (# Measure ...):
      - BrainSegVol               -> Volumen cerebral total
      - eTIV                      -> Volumen total intracraneal (también lo leemos aparte)
      - CerebralWhiteMatterVol    -> Sustancia blanca total
      - TotalGrayVol              -> Sustancia gris total
      - SubCortGrayVol            -> Sustancia gris profunda
      - lhCortexVol               -> Sustancia gris corteza izquierda
      - rhCortexVol               -> Sustancia gris corteza derecha
    """
    out = {}
    aseg_stats = os.path.join(stats_dir, "aseg.stats")
    if not os.path.exists(aseg_stats):
        _warn(f"No se encontró '{aseg_stats}'. Métricas globales quedarán como NaN si no derivan de otras fuentes.")
        return out
    m = _parse_header_measures(aseg_stats)
    keys = [
        "BrainSegVol", "eTIV", "CerebralWhiteMatterVol", "TotalGrayVol",
        "SubCortGrayVol", "lhCortexVol", "rhCortexVol"
    ]
    for k in keys:
        if k in m:
            out[k] = float(m[k])
    return out

# ------------------------------- Lóbulos (corteza) y subcorteza ----------------------------------

_LOBES: Dict[str, List[str]] = {
    'Frontal': [
        'caudalanteriorcingulate','caudalmiddlefrontal','lateralorbitofrontal',
        'medialorbitofrontal','parsopercularis','parsorbitalis','parstriangularis',
        'precentral','superiorfrontal','rostralanteriorcingulate','rostralmiddlefrontal','paracentral'
    ],
    'Parietal': ['superiorparietal','inferiorparietal','supramarginal','postcentral','precuneus'],
    'Temporal': ['superiortemporal','middletemporal','inferiortemporal','fusiform','transversetemporal','parahippocampal'],
    'Occipital': ['lateraloccipital','cuneus','pericalcarine','lingual'],
}

def _lobar_volumes_mm3(stats_dir: str) -> Tuple[Dict[str, float], Dict[str, float]]:
    """
    Suma GrayVol por lóbulo desde:
      - lh_aparc.DKTatlas.mapped_volume_stats.txt (LH)
      - rh_aparc.DKTatlas.mapped_volume_stats.txt (RH)
    Devuelve:
      lobes_total_mm3: {'Frontal': total, ...}
      lobes_hemi_mm3:  {'Frontal_LH': v, 'Frontal_RH': v, ...}
    """
    lh_path = os.path.join(stats_dir, "lh_aparc.DKTatlas.mapped_volume_stats.txt")
    rh_path = os.path.join(stats_dir, "rh_aparc.DKTatlas.mapped_volume_stats.txt")
    s_lh = _read_transposed_series(lh_path)  # e.g., 'lh_cuneus_volume'
    s_rh = _read_transposed_series(rh_path)  # e.g., 'rh_cuneus_volume'

    def _get_from_series(ser: pd.Series, hemi: str, roi: str) -> float:
        # Patrones robustos
        for pat in (f"{hemi}_{roi}_volume", f"{hemi}.{roi}.volume", f"{hemi}-{roi}-volume"):
            if pat in ser.index:
                return float(ser[pat])
        # última chance: solo roi (si archivo separado por hemi)
        if roi in ser.index:
            return float(ser[roi])
        _warn(f"ROI ausente para lóbulo: {hemi} {roi}")
        return float("nan")

    lobes_total_mm3, lobes_hemi_mm3 = {}, {}
    for lobe, rois in _LOBES.items():
        lh = np.nansum([_get_from_series(s_lh, "lh", r) for r in rois])
        rh = np.nansum([_get_from_series(s_rh, "rh", r) for r in rois])
        lobes_total_mm3[lobe] = float(lh + rh)
        lobes_hemi_mm3[f"{lobe}_LH"] = float(lh)
        lobes_hemi_mm3[f"{lobe}_RH"] = float(rh)
    return lobes_total_mm3, lobes_hemi_mm3

def _subcortical_volumes_mm3(stats_dir: str) -> Tuple[Dict[str, float], Dict[str, float]]:
    """
    Desde aseg_stats_cm3.txt (cm3, transpuesto) -> mm3
    Devuelve:
      subcort_total_mm3: {'Hipocampo': total, 'Amígdala': total, 'Cuerpo Calloso': total, ...,
                          'Ventrículo lateral izquierdo': vol, 'Ventrículo lateral derecho': vol,
                          'Ventrículos Laterales': total}
      subcort_hemi_mm3:  {'Hipocampo_LH': v, 'Hipocampo_RH': v, 'Amígdala_LH': ..., 'Tálamo_LH': ...}
    """
    path = os.path.join(stats_dir, "aseg_stats_cm3.txt")
    s_cm3 = _read_transposed_series(path)  # ROI -> cm3
    get = lambda roi: float(s_cm3.get(roi, np.nan)) if pd.notna(s_cm3.get(roi, np.nan)) else np.nan

    def thal(side: str) -> float:
        v = get(f"{side}-Thalamus-Proper")
        if not np.isfinite(v):
            v = get(f"{side}-Thalamus")
        return v

    # En cm3
    lh_hip = get("Left-Hippocampus")
    rh_hip = get("Right-Hippocampus")
    lh_amg = get("Left-Amygdala")
    rh_amg = get("Right-Amygdala")
    lh_tha = thal("Left")
    rh_tha = thal("Right")
    lh_cau = get("Left-Caudate")
    rh_cau = get("Right-Caudate")
    lh_put = get("Left-Putamen")
    rh_put = get("Right-Putamen")
    lh_acc = get("Left-Accumbens-area")
    rh_acc = get("Right-Accumbens-area")

    cc_parts = ["CC_Posterior","CC_Mid_Posterior","CC_Central","CC_Mid_Anterior","CC_Anterior"]
    cc_cm3 = np.nansum([get(x) for x in cc_parts])

    lv_l_cm3 = get("Left-Lateral-Ventricle")
    lv_r_cm3 = get("Right-Lateral-Ventricle")

    # a mm3
    to_mm3 = lambda x: float(x * 1000.0) if np.isfinite(x) else float("nan")

    subcort_total_mm3 = {
        "Hipocampo": to_mm3(lh_hip + rh_hip),
        "Amígdala": to_mm3(lh_amg + rh_amg),
        "Cuerpo Calloso": to_mm3(cc_cm3),
        "Ventrículo lateral izquierdo": to_mm3(lv_l_cm3),
        "Ventrículo lateral derecho": to_mm3(lv_r_cm3),
        "Ventrículos Laterales": to_mm3((lv_l_cm3 or 0) + (lv_r_cm3 or 0)),
        "Ganglios basales: estriado": to_mm3((lh_cau or 0)+(rh_cau or 0)+(lh_put or 0)+(rh_put or 0)+(lh_acc or 0)+(rh_acc or 0)),
        "Ganglios basales: tálamo": to_mm3((lh_tha or 0) + (rh_tha or 0)),
    }
    subcort_hemi_mm3 = {
        "Hipocampo_LH": to_mm3(lh_hip),   "Hipocampo_RH": to_mm3(rh_hip),
        "Amígdala_LH": to_mm3(lh_amg),    "Amígdala_RH": to_mm3(rh_amg),
        "Estriado_LH": to_mm3((lh_cau or 0)+(lh_put or 0)+(lh_acc or 0)),
        "Estriado_RH": to_mm3((rh_cau or 0)+(rh_put or 0)+(rh_acc or 0)),
        "Tálamo_LH": to_mm3(lh_tha),      "Tálamo_RH": to_mm3(rh_tha),
        "Ventrículos_LH": to_mm3(lv_l_cm3), "Ventrículos_RH": to_mm3(lv_r_cm3),
    }
    return subcort_total_mm3, subcort_hemi_mm3

# ------------------------------- Construcción de tablas finales ----------------------------------
def _build_comparison_table(
    stats_dir: str,
    base_control_path: str
) -> Tuple[pd.DataFrame, Dict[str, float], Dict[str, float], Dict[str, float]]:
    """
    Arma df_comparacion (sin estilos) y devuelve también diccionarios útiles para asimetrías.
    - Volrel% (sujeto) se calcula SIEMPRE después de armar df_subj, a partir de Volumen mm3 (sujeto) y eTIV.
    - Espesores: Volrel% (sujeto) = NaN (como pediste).
    """

    # ------------- eTIV (mm3) -------------
    etiv = _get_etiv_mm3(stats_dir)  # si falta -> NaN (warning interno)
    if not np.isfinite(etiv) or etiv <= 0:
        print("[WARN] eTIV no disponible o inválido; Volrel% (sujeto) quedará NaN.")

    # ------------- Lóbulos (GrayVol) -------------
    lobes_total_mm3, lobes_hemi_mm3 = _lobar_volumes_mm3(stats_dir)

    # ------------- Subcorteza -------------
    subc_total_mm3, subc_hemi_mm3 = _subcortical_volumes_mm3(stats_dir)

    # ------------- Globales desde aseg.stats -------------
    globals_meas = _get_global_measures_mm3(stats_dir)
    # globals_meas keys esperados: BrainSegVol, eTIV, CerebralWhiteMatterVol,
    # TotalGrayVol, SubCortGrayVol, lhCortexVol, rhCortexVol, ...

    # ------------- Espesores (mm) -------------
    thick_lh, thick_rh, thick_mean = _get_thickness_mm(stats_dir)

    # ============ Construir filas del SUJETO (primero SIEMPRE en mm3/valor crudo) ============
    filas = []

    # Espesores (guardamos valor del espesor en la columna de "Volumen mm3 (sujeto)" para no romper estructura;
    # Volrel% (sujeto) se dejará en NaN luego)
    filas += [
        ("Espesor cortical derecho (mm)", float("nan"), thick_rh),
        ("Espesor cortical izquierdo (mm)", float("nan"), thick_lh),
        ("Espesor cortical promedio (mm)", float("nan"), thick_mean),
    ]

    # Lóbulos (usar suma total por lóbulo)
    for lobe in ["Frontal", "Occipital", "Parietal", "Temporal"]:
        vol_mm3 = float(lobes_total_mm3.get(lobe, float("nan")))
        filas.append((lobe, float("nan"), vol_mm3))  # volrel se calcula luego

    # Subcorteza y globales
    nombres = [
        "Amígdala", "Cuerpo Calloso", "Ganglios basales: estriado", "Ganglios basales: tálamo",
        "Hipocampo", "Sustancia blanca total", "Sustancia gris corteza derecha",
        "Sustancia gris corteza izquierda", "Sustancia gris profunda", "Sustancia gris total",
        "Ventrículo lateral derecho", "Ventrículo lateral izquierdo", "Ventrículos Laterales",
        "Volumen cerebral total", "Volumen total intracraneal"
    ]
    for name in nombres:
        if name == "Sustancia blanca total":
            vol_mm3 = float(globals_meas.get("CerebralWhiteMatterVol", float("nan")))
        elif name == "Sustancia gris total":
            vol_mm3 = float(globals_meas.get("TotalGrayVol", float("nan")))
        elif name == "Sustancia gris profunda":
            vol_mm3 = float(globals_meas.get("SubCortGrayVol", float("nan")))
        elif name == "Sustancia gris corteza izquierda":
            vol_mm3 = float(globals_meas.get("lhCortexVol", float("nan")))
        elif name == "Sustancia gris corteza derecha":
            vol_mm3 = float(globals_meas.get("rhCortexVol", float("nan")))
        elif name == "Volumen cerebral total":
            vol_mm3 = float(globals_meas.get("BrainSegVol", float("nan")))
        elif name == "Volumen total intracraneal":
            # preferimos el eTIV del header si existe
            vol_mm3 = float(globals_meas.get("eTIV", etiv))
        else:
            vol_mm3 = float(subc_total_mm3.get(name, float("nan")))

        filas.append((name, float("nan"), vol_mm3))

    # DataFrame del sujeto (crudo: primero definimos mm3, luego derivamos Volrel%)
    df_subj = pd.DataFrame(filas, columns=["Measure:GrayVol", "Volrel% (sujeto)", "Volumen mm3 (sujeto)"])

    # ============ Calcular Volrel% (sujeto) AHORA ============
    def _calc_volrel(row):
        name = str(row["Measure:GrayVol"]).lower()
        if name.startswith("espesor cortical"):
            return float("nan")  # espesores sin volrel
        v = pd.to_numeric(row["Volumen mm3 (sujeto)"], errors="coerce")
        if not np.isfinite(v) or not np.isfinite(etiv) or etiv <= 0:
            return float("nan")
        return float((v / etiv) * 100.0)

    df_subj["Volrel% (sujeto)"] = df_subj.apply(_calc_volrel, axis=1)

    # ============ Cargar base control y unir ============
    df_ctrl = pd.read_excel(base_control_path, sheet_name="metricas", engine="openpyxl")

    # Normalizar llaves para merge robusto
    df_ctrl["_key"] = df_ctrl["Measure:GrayVol"].map(_normalize_key)
    df_subj["_key"] = df_subj["Measure:GrayVol"].map(_normalize_key)

    cols_needed = [
        "Measure:GrayVol","Volumen mm3","Mediana","MAD","Volrel%","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto","valores_muestreo","_key"
    ]
    missing = [c for c in cols_needed if c not in df_ctrl.columns]
    if missing:
        raise ValueError(f"Faltan columnas en hoja 'metricas': {missing}")

    df_merged = pd.merge(
        df_subj,
        df_ctrl[cols_needed],
        on="_key",
        how="left",
        suffixes=("_suj","_ctrl")
    )

    # Asegurar que nuestras columnas clave no se contaminen tras el merge
    # (las del sujeto mantienen su nombre exacto)
    # Measure final: el texto del Excel (para mantener acentos/casos idénticos)
    df_merged["Measure:GrayVol"] = df_merged["Measure:GrayVol_ctrl"]

    # ============ Percentil y Z (sólo espesores) ============
    percentiles, z_vals, dentro_35 = [], [], []
    for _, row in df_merged.iterrows():
        name = row["Measure:GrayVol"]
        valores = _parse_valores_muestreo_to_sorted(row["valores_muestreo"])
        if str(name).lower().startswith("espesor cortical"):
            # valor x = espesor (mm) sujeto
            if "derecho" in _normalize_key(name):
                x = thick_rh
            elif "izquierdo" in _normalize_key(name):
                x = thick_lh
            else:
                x = thick_mean
            p = _percentil_lineal(valores, x)
            z = _z_robusto(x, row["Mediana"], row["MAD"])
            dentro = "✓" if (np.isfinite(z) and abs(z) <= 3.5) else ""
        else:
            # volúmenes: percentil contra Volrel% (sujeto)
            x = row["Volrel% (sujeto)"]
            p = _percentil_lineal(valores, x)
            z = float("nan")
            dentro = ""
        percentiles.append(p)
        z_vals.append(z)
        dentro_35.append(dentro)

    df_merged["Percentil (sujeto)"] = percentiles
    df_merged["Z_sujeto"] = z_vals
    df_merged["Dentro_de_Umbral_±3.5"] = dentro_35

    # ============ Flag (IC99 en Volrel%) ============
    flags = []
    for _, row in df_merged.iterrows():
        name = row["Measure:GrayVol"]
        if str(name).lower().startswith("espesor cortical"):
            flags.append("")  # espesores sin flag
        else:
            v = pd.to_numeric(row["Volrel% (sujeto)"], errors="coerce")
            b99 = pd.to_numeric(row["IC_99%_Bajo"], errors="coerce")
            a99 = pd.to_numeric(row["IC_99%_Alto"], errors="coerce")
            if not np.isfinite(v) or not np.isfinite(b99) or not np.isfinite(a99):
                flags.append("")  # sin datos suficientes
            else:
                flags.append("Dentro de rango" if (b99 <= v <= a99) else "Fuera de rango")
    df_merged["Flag"] = flags

    # ============ Selección y orden final ============
    keep_cols = [
        "Measure:GrayVol",
        "Volumen mm3 (sujeto)",
        "Volrel% (sujeto)",
        "Mediana","Volrel%","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto",
        "Percentil (sujeto)","Flag","Z_sujeto","Dentro_de_Umbral_±3.5"
    ]
    df_final = df_merged[keep_cols].copy()

    # Orden solicitado
    orden = [
        "Espesor cortical derecho (mm)",
        "Espesor cortical izquierdo (mm)",
        "Espesor cortical promedio (mm)",
        "Frontal","Occipital","Parietal","Temporal",
        "Amígdala","Cuerpo Calloso","Ganglios basales: estriado","Ganglios basales: tálamo",
        "Hipocampo","Sustancia blanca total","Sustancia gris corteza derecha","Sustancia gris corteza izquierda",
        "Sustancia gris profunda","Sustancia gris total","Ventrículo lateral derecho","Ventrículo lateral izquierdo",
        "Ventrículos Laterales","Volumen cerebral total","Volumen total intracraneal"
    ]
    cat = pd.Categorical(df_final["Measure:GrayVol"], categories=orden, ordered=True)
    df_final = df_final.assign(_ord=cat).sort_values("_ord").drop(columns=["_ord"])

    # Devolver también diccionarios útiles
    return df_final, lobes_hemi_mm3, subc_hemi_mm3, {"thick_lh": thick_lh, "thick_rh": thick_rh}

def _calc_li(left: float, right: float) -> float:
    """LI% = ((L - R)/((L+R)/2))*100; NaN si faltantes o suma=0."""
    if not (np.isfinite(left) and np.isfinite(right)) or (left + right) == 0:
        return float("nan")
    return ( (left - right) / ( (left + right) / 2.0 ) ) * 100.0

def _build_asymmetry_table(
    base_control_path: str,
    lobes_hemi_mm3: Dict[str, float],
    subc_hemi_mm3: Dict[str, float],
    globals_meas: Dict[str, float],
    thick_lh: float,
    thick_rh: float
) -> pd.DataFrame:
    """
    Construye df_asimetrias usando pares LH/RH y la hoja 'Asimetrias' (rangos).
    - La hoja 'Asimetrias' del Excel base tiene columnas:
      ['Measure:GrayVol','Mediana','IC_95%_Bajo','IC_95%_Alto','IC_99%_Bajo','IC_99%_Alto','rango normal']
    - Calcula LI% del sujeto donde aplica (espesor, lóbulos, pares subcorticales, Ventrículos Laterales,
      y también para 'Sustancia gris corteza derecha/izquierda' usando lhCortexVol/rhCortexVol).
    - Para filas sin par (p.ej. Cuerpo Calloso, Ventrículo lateral derecho/izquierdo, globales) deja LI=NaN.
    - Flag: 'Dentro de rango' si LI ∈ [IC_99%_Bajo, IC_99%_Alto], sino 'Fuera de rango'.
      Si faltan LI o límites, deja 'Dentro de rango' por defecto.
    """
    def _safe_get(d, k):
        return float(d.get(k, float("nan")))

    # LI por categorías
    li_rows = []

    # Espesores (mismo LI replicado en las tres filas, como en tu ejemplo)
    li_thick = _calc_li(thick_lh, thick_rh)
    li_rows.append(("Espesor cortical derecho (mm)", li_thick))
    li_rows.append(("Espesor cortical izquierdo (mm)", li_thick))
    li_rows.append(("Espesor cortical promedio (mm)", li_thick))

    # Lóbulos (mm3)
    for lobe in ["Frontal","Occipital","Parietal","Temporal"]:
        L = _safe_get(lobes_hemi_mm3, f"{lobe}_LH")
        R = _safe_get(lobes_hemi_mm3, f"{lobe}_RH")
        li_rows.append((lobe, _calc_li(L, R)))

    # Subcorteza pares (mm3)
    # Estriado
    L = _safe_get(subc_hemi_mm3, "Estriado_LH")
    R = _safe_get(subc_hemi_mm3, "Estriado_RH")
    li_rows.append(("Ganglios basales: estriado", _calc_li(L, R)))
    # Tálamo
    L = _safe_get(subc_hemi_mm3, "Tálamo_LH")
    R = _safe_get(subc_hemi_mm3, "Tálamo_RH")
    li_rows.append(("Ganglios basales: tálamo", _calc_li(L, R)))
    # Hipocampo
    L = _safe_get(subc_hemi_mm3, "Hipocampo_LH")
    R = _safe_get(subc_hemi_mm3, "Hipocampo_RH")
    li_rows.append(("Hipocampo", _calc_li(L, R)))
    # Amígdala
    L = _safe_get(subc_hemi_mm3, "Amígdala_LH")
    R = _safe_get(subc_hemi_mm3, "Amígdala_RH")
    li_rows.append(("Amígdala", _calc_li(L, R)))
    # Ventrículos Laterales (suma L/R)
    L = _safe_get(subc_hemi_mm3, "Ventrículos_LH")
    R = _safe_get(subc_hemi_mm3, "Ventrículos_RH")
    li_rows.append(("Ventrículos Laterales", _calc_li(L, R)))

    # Sustancia gris corteza derecha/izquierda -> LI desde lhCortexVol/rhCortexVol
    lh_ctx = float(globals_meas.get("lhCortexVol", float("nan")))
    rh_ctx = float(globals_meas.get("rhCortexVol", float("nan")))
    li_ctx = _calc_li(lh_ctx, rh_ctx)
    li_rows.append(("Sustancia gris corteza derecha", li_ctx))
    li_rows.append(("Sustancia gris corteza izquierda", li_ctx))

    # Filas sin LI (se dejan en NaN y Flag por defecto 'Dentro de rango')
    for name in [
        "Cuerpo Calloso",
        "Sustancia blanca total",
        "Sustancia gris profunda",
        "Sustancia gris total",
        "Ventrículo lateral derecho",
        "Ventrículo lateral izquierdo",
        "Volumen cerebral total",
        "Volumen total intracraneal",
    ]:
        li_rows.append((name, float("nan")))

    df_li = pd.DataFrame(li_rows, columns=["Measure:GrayVol","LI% (Volrel)"])

    # Unir con hoja Asimetrias (base)
    df_ref = pd.read_excel(base_control_path, sheet_name="Asimetrias", engine="openpyxl")
    need_cols = ["Measure:GrayVol","Mediana","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto","rango normal"]
    miss = [c for c in need_cols if c not in df_ref.columns]
    if miss:
        raise ValueError(f"Faltan columnas en hoja 'Asimetrias': {miss}")

    df_ref["_key"] = df_ref["Measure:GrayVol"].map(_normalize_key)
    df_li["_key"]  = df_li["Measure:GrayVol"].map(_normalize_key)

    df_asim = pd.merge(
        df_li,
        df_ref[need_cols + ["_key"]],
        on="_key",
        how="left",
        suffixes=("_suj","_ref")
    )

    # Flag del sujeto (IC99). Si falta algo -> 'Dentro de rango' (como en tu ejemplo para filas sin LI)
    flags = []
    for _, row in df_asim.iterrows():
        v   = row["LI% (Volrel)"]
        b99 = row["IC_99%_Bajo"]
        a99 = row["IC_99%_Alto"]
        if pd.isna(v) or pd.isna(b99) or pd.isna(a99):
            flags.append("Dentro de rango")
        else:
            flags.append("Dentro de rango" if (b99 <= v <= a99) else "Fuera de rango")
    df_asim["Flag"] = flags

    # Selección/rename de columnas exactamente como pediste
    df_asim["Measure:GrayVol"] = df_asim["Measure:GrayVol_ref"]
    keep = [
        "Measure:GrayVol",
        "LI% (Volrel)",
        "Mediana","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto",
        "Flag","rango normal"
    ]
    df_asim = df_asim[keep]

    # Orden final de filas (tu orden)
    orden = [
        "Espesor cortical derecho (mm)","Espesor cortical izquierdo (mm)","Espesor cortical promedio (mm)",
        "Frontal","Occipital","Parietal","Temporal",
        "Amígdala","Cuerpo Calloso","Ganglios basales: estriado","Ganglios basales: tálamo",
        "Hipocampo","Sustancia blanca total","Sustancia gris corteza derecha","Sustancia gris corteza izquierda",
        "Sustancia gris profunda","Sustancia gris total",
        "Ventrículo lateral derecho","Ventrículo lateral izquierdo","Ventrículos Laterales",
        "Volumen cerebral total","Volumen total intracraneal"
    ]
    cat = pd.Categorical(df_asim["Measure:GrayVol"], categories=orden, ordered=True)
    df_asim = df_asim.assign(_ord=cat).sort_values("_ord").drop(columns=["_ord"])

    # Redondeo a 2 decimales (valores numéricos)
    for col in ["LI% (Volrel)","Mediana","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto"]:
        if col in df_asim.columns:
            df_asim[col] = pd.to_numeric(df_asim[col], errors="coerce").round(2)

    return df_asim


# ------------------------------- Export a Excel con estilos --------------------------------------

def _style_and_export(
    out_path: str,
    df_vol: pd.DataFrame,
    df_asim: pd.DataFrame
):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_vol.to_excel(writer, sheet_name="Volumenes", index=False)
        df_asim.to_excel(writer, sheet_name="Asimetrias", index=False)

        wb = writer.book
        normal_style = NamedStyle(name="arial_10", font=Font(name="Arial", size=10))
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Volúmenes: resaltar fuera IC95 (negrita) y fuera IC99 (naranja) usando Volrel% (sujeto)
        ws = writer.sheets["Volumenes"]
        # aplicar estilo base
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.style = normal_style

        # columnas por nombre
        hdr = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        col_vrel = hdr.get("Volrel% (sujeto)")
        col_ic95_b = hdr.get("IC_95%_Bajo")
        col_ic95_a = hdr.get("IC_95%_Alto")
        col_ic99_b = hdr.get("IC_99%_Bajo")
        col_ic99_a = hdr.get("IC_99%_Alto")
        col_name  = hdr.get("Measure:GrayVol")

        for r in range(2, ws.max_row+1):
            name = ws.cell(row=r, column=col_name).value or ""
            name_norm = _normalize_key(str(name))
            # no estilamos filas de "espesor cortical"
            if name_norm.startswith("espesor cortical"):
                continue
            try:
                v  = float(ws.cell(row=r, column=col_vrel).value)
                i95b = float(ws.cell(row=r, column=col_ic95_b).value)
                i95a = float(ws.cell(row=r, column=col_ic95_a).value)
                i99b = float(ws.cell(row=r, column=col_ic99_b).value)
                i99a = float(ws.cell(row=r, column=col_ic99_a).value)
            except Exception:
                continue

            # fuera de IC95 -> negrita
            if not (i95b <= v <= i95a):
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(name="Arial", size=10, bold=True)
            # fuera de IC99 -> naranja
            if not (i99b <= v <= i99a):
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = orange_fill

        # Asimetrías: naranja si fuera de IC99
        ws2 = writer.sheets["Asimetrias"]
        hdr2 = {cell.value: idx+1 for idx, cell in enumerate(ws2[1])}
        col_li   = hdr2.get("LI% (Volrel)")
        col_a99b = hdr2.get("IC_99%_Bajo")
        col_a99a = hdr2.get("IC_99%_Alto")
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
            for cell in row:
                cell.style = normal_style
        for r in range(2, ws2.max_row+1):
            try:
                v  = ws2.cell(row=r, column=col_li).value
                b  = ws2.cell(row=r, column=col_a99b).value
                a  = ws2.cell(row=r, column=col_a99a).value
                if v is None or b is None or a is None:
                    continue
                v, b, a = float(v), float(b), float(a)
            except Exception:
                continue
            if not (b <= v <= a):
                for c in range(1, ws2.max_column+1):
                    ws2.cell(row=r, column=c).fill = orange_fill

# ------------------------------- Función principal pública ---------------------------------------

def comparar_morfometria_y_exportar(stats_dir: str, base_control_path: str, output_excel_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Orquesta todo:
      1) Construye tabla de comparación con base 'metricas' (volúmenes y espesores).
      2) Construye tabla de asimetrías con base 'Asimetrias'.
      3) Exporta Excel con dos hojas estiladas.
    Retorna (df_comparacion, df_asimetrias).
    """
    # Tabla de comparación y materiales para asimetrías
    df_comp, lobes_hemi_mm3, subc_hemi_mm3, thick_dict = _build_comparison_table(stats_dir, base_control_path)

    # Globales (para consistencia; aunque no se usan aquí explícitamente)
    globals_meas = _get_global_measures_mm3(stats_dir)

    # Asimetrías
    df_asim = _build_asymmetry_table(
        base_control_path=base_control_path,
        lobes_hemi_mm3=lobes_hemi_mm3,
        subc_hemi_mm3=subc_hemi_mm3,
        globals_meas=globals_meas,
        thick_lh=thick_dict["thick_lh"],
        thick_rh=thick_dict["thick_rh"]
    )

    # Redondeo final a 2 decimales (según requerimiento) manteniendo tipo numérico
    def _round_cols(df: pd.DataFrame, cols: List[str], ndigits: int = 2):
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").round(ndigits)

    _round_cols(df_comp, ["Volumen mm3 (sujeto)","Volrel% (sujeto)","Mediana","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto","Percentil (sujeto)","Z_sujeto"], 2)
    _round_cols(df_asim, ["LI% (Volrel)","Mediana","IC_95%_Bajo","IC_95%_Alto","IC_99%_Bajo","IC_99%_Alto"], 2)

    # Export estilado
    _style_and_export(output_excel_path, df_comp, df_asim)

    print(f"\nExcel exportado en: {output_excel_path}")

    return df_comp, df_asim